from __future__ import annotations

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import pandas as pd
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
except Exception:  # pragma: no cover
    sync_playwright = None
    PlaywrightTimeoutError = Exception


IBD_LISTS = {
    "IBD 50": "https://research.investors.com/stock-lists/ibd-50/",
    "IBD Big Cap 20": "https://research.investors.com/stock-lists/big-cap-20/",
    "IBD Sector Leaders": "https://research.investors.com/stock-lists/sector-leaders",
    "Stock Spotlight": "https://research.investors.com/stock-lists/stock-spotlight/",
    "IPO Leaders": "https://research.investors.com/stock-lists/ipo-leaders/",
    "New Highs": "https://research.investors.com/stock-lists/new-highs/",
    "Relative Strength at New High": "https://research.investors.com/stock-lists/relative-strength-at-new-high/",
    "Global Leaders": "https://research.investors.com/stock-lists/global-leaders/",
    "Rising Profit Estimates": "https://research.investors.com/stock-lists/rising-profit-estimates/",
    "Stocks that Funds are Buying": "https://research.investors.com/stock-lists/stocks-that-funds-are-buying/",
    "Your Weekly Review": "https://research.investors.com/stock-lists/your-weekly-review/",
}

DEFAULT_DOWNLOAD_DIR = Path("downloads")
DEFAULT_OUTPUT = Path("ibd_current_lists_compiled.xlsx")


@dataclass
class ExportedList:
    list_name: str
    url: str
    pulled_at_utc: str
    file_path: Path


class CompileRequest(BaseModel):
    download_dir: str = Field(default="downloads")
    output_xlsx: str = Field(default="ibd_current_lists_compiled.xlsx")
    manual_login: bool = Field(default=True)
    headless: bool = Field(default=False)
    email: str | None = None
    password: str | None = None


app = FastAPI(title="IBD List Compiler", version="1.0.0")


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return value.lower() or "sheet"


def clean_header(value: object, index: int) -> str:
    text = str(value).strip() if value is not None else ""
    text = re.sub(r"\s+", " ", text)
    if not text or text.lower().startswith("unnamed:"):
        return f"column_{index + 1}"
    return text


def dedupe_headers(headers: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    output: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        output.append(header if count == 0 else f"{header}_{count + 1}")
    return output


TICKER_CANDIDATES = [
    "ticker",
    "symbol",
    "stock symbol",
    "stock",
    "company symbol",
]
COMPANY_CANDIDATES = [
    "company",
    "company name",
    "name",
    "stock name",
]


def find_best_column(columns: Iterable[str], candidates: list[str]) -> str | None:
    normalized = {c.lower().strip(): c for c in columns}
    for cand in candidates:
        if cand in normalized:
            return normalized[cand]
    for col in columns:
        lc = col.lower().strip()
        if any(cand in lc for cand in candidates):
            return col
    return None


def read_export_file(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    elif path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Unsupported export file type: {path.name}")

    headers = dedupe_headers([clean_header(col, idx) for idx, col in enumerate(df.columns)])
    df.columns = headers
    df = df.dropna(how="all")
    df = df.loc[:, ~(df.columns.str.startswith("column_") & df.isna().all())]
    df = df.reset_index(drop=True)
    return df


def autosize_worksheet(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max(max_len + 2, 10), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_worksheet(ws, table_name: str) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name[:255], ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    autosize_worksheet(ws)
    ws.row_dimensions[1].height = 24


def build_workbook(exports: list[ExportedList], output_path: Path) -> Path:
    if not exports:
        raise ValueError("No export files found to compile.")

    all_frames: list[pd.DataFrame] = []
    wb = Workbook()
    wb.remove(wb.active)

    metadata_rows = []

    for idx, item in enumerate(exports, start=1):
        df = read_export_file(item.file_path)
        ticker_col = find_best_column(df.columns, TICKER_CANDIDATES)
        company_col = find_best_column(df.columns, COMPANY_CANDIDATES)

        df.insert(0, "List Name", item.list_name)
        df.insert(1, "List URL", item.url)
        df.insert(2, "Pulled At UTC", item.pulled_at_utc)
        df.insert(3, "Source File", item.file_path.name)
        if ticker_col:
            df.insert(4, "Ticker Normalized", df[ticker_col].astype(str).str.strip().str.upper())
        else:
            df.insert(4, "Ticker Normalized", "")
        if company_col:
            df.insert(5, "Company Normalized", df[company_col].astype(str).str.strip())
        else:
            df.insert(5, "Company Normalized", "")

        all_frames.append(df)

        ws_name = re.sub(r"[\\/*?:\[\]]", "", item.list_name)[:31]
        ws = wb.create_sheet(title=ws_name)
        ws.append(df.columns.tolist())
        for row in df.fillna("").itertuples(index=False, name=None):
            ws.append(list(row))
        style_worksheet(ws, f"tbl_{slugify(ws_name)}")

        metadata_rows.append([
            idx,
            item.list_name,
            item.url,
            item.pulled_at_utc,
            item.file_path.name,
            len(df),
        ])

    combined = pd.concat(all_frames, ignore_index=True, sort=False).fillna("")
    combined_cols = list(combined.columns)
    if "Ticker Normalized" in combined_cols:
        sort_cols = [c for c in ["Ticker Normalized", "List Name"] if c in combined.columns]
        if sort_cols:
            combined = combined.sort_values(sort_cols, kind="stable")

    ws_all = wb.create_sheet(title="All Lists")
    ws_all.append(combined.columns.tolist())
    for row in combined.itertuples(index=False, name=None):
        ws_all.append(list(row))
    style_worksheet(ws_all, "tbl_all_lists")

    unique_cols = [c for c in ["Ticker Normalized", "Company Normalized"] if c in combined.columns]
    if unique_cols:
        summary = (
            combined.groupby(unique_cols, dropna=False)
            .agg(
                Lists_Count=("List Name", "nunique"),
                Lists=("List Name", lambda s: ", ".join(sorted(set(map(str, s))))),
            )
            .reset_index()
            .sort_values(["Lists_Count", unique_cols[0]], ascending=[False, True], kind="stable")
        )
        ws_summary = wb.create_sheet(title="Unique Tickers")
        ws_summary.append(summary.columns.tolist())
        for row in summary.fillna("").itertuples(index=False, name=None):
            ws_summary.append(list(row))
        style_worksheet(ws_summary, "tbl_unique_tickers")

    ws_meta = wb.create_sheet(title="Run Metadata")
    ws_meta.append(["#", "List Name", "List URL", "Pulled At UTC", "Source File", "Row Count"])
    for row in metadata_rows:
        ws_meta.append(row)
    style_worksheet(ws_meta, "tbl_run_metadata")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def prompt_manual_login(page, first_url: str) -> None:
    print("Opening IBD in a browser window.")
    print("Log in manually, complete any MFA/challenges, then press Enter here.")
    page.goto(first_url, wait_until="domcontentloaded")
    input("Press Enter after you can see a logged-in IBD page... ")


def fill_login_if_possible(page, email: str | None, password: str | None) -> bool:
    if not email or not password:
        return False

    page.goto("https://www.investors.com/", wait_until="domcontentloaded")
    time.sleep(2)
    try_selectors = [
        "text=Log In",
        "text=Sign In",
        "a[href*='login']",
        "button:has-text('Log In')",
    ]
    clicked = False
    for selector in try_selectors:
        try:
            page.locator(selector).first.click(timeout=3000)
            clicked = True
            break
        except Exception:
            continue

    if not clicked:
        return False

    time.sleep(2)
    email_selectors = ["input[type='email']", "input[name='email']", "input[name='username']"]
    password_selectors = ["input[type='password']", "input[name='password']"]

    for selector in email_selectors:
        try:
            page.locator(selector).first.fill(email, timeout=3000)
            break
        except Exception:
            pass
    else:
        return False

    for selector in password_selectors:
        try:
            page.locator(selector).first.fill(password, timeout=3000)
            break
        except Exception:
            pass
    else:
        return False

    for selector in ["button[type='submit']", "button:has-text('Log In')", "button:has-text('Sign In')"]:
        try:
            page.locator(selector).first.click(timeout=3000)
            time.sleep(5)
            return True
        except Exception:
            continue
    return False


def click_export_and_download(page, list_name: str, url: str, download_dir: Path) -> Path:
    page.goto(url, wait_until="domcontentloaded")
    page.wait_for_timeout(3000)

    selectors = [
        "text=To Excel",
        "text=Export",
        "button:has-text('Export')",
        "button:has-text('To Excel')",
        "[aria-label*='Export']",
        "[title*='Export']",
    ]

    last_error: Exception | None = None
    for selector in selectors:
        try:
            with page.expect_download(timeout=15000) as download_info:
                page.locator(selector).first.click(timeout=4000)
            download = download_info.value
            filename = download.suggested_filename
            ext = Path(filename).suffix or ".xlsx"
            out_path = download_dir / f"{slugify(list_name)}{ext}"
            download.save_as(str(out_path))
            return out_path
        except Exception as exc:
            last_error = exc
            continue

    raise RuntimeError(f"Could not export {list_name} from {url}. Last error: {last_error}")


def download_all_lists(
    download_dir: Path,
    headless: bool = False,
    manual_login: bool = True,
    email: str | None = None,
    password: str | None = None,
) -> list[ExportedList]:
    if sync_playwright is None:
        raise RuntimeError("Playwright is not installed. Install dependencies first.")

    download_dir.mkdir(parents=True, exist_ok=True)
    pulled_at_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    results: list[ExportedList] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        first_url = next(iter(IBD_LISTS.values()))
        logged_in = fill_login_if_possible(page, email, password)
        if manual_login or not logged_in:
            prompt_manual_login(page, first_url)

        for list_name, url in IBD_LISTS.items():
            print(f"Exporting {list_name}...")
            file_path = click_export_and_download(page, list_name, url, download_dir)
            results.append(
                ExportedList(
                    list_name=list_name,
                    url=url,
                    pulled_at_utc=pulled_at_utc,
                    file_path=file_path,
                )
            )
        context.close()
        browser.close()

    return results


def discover_downloads(download_dir: Path) -> list[ExportedList]:
    if not download_dir.exists():
        raise FileNotFoundError(f"Download directory not found: {download_dir}")

    pulled_at_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    discovered: list[ExportedList] = []

    slug_to_name = {slugify(name): name for name in IBD_LISTS}
    name_to_url = IBD_LISTS.copy()

    for path in sorted(download_dir.iterdir()):
        if path.suffix.lower() not in {".csv", ".xlsx", ".xls", ".xlsm"}:
            continue

        slug = slugify(path.stem)
        list_name = slug_to_name.get(slug)
        if not list_name:
            for maybe_slug, maybe_name in slug_to_name.items():
                if maybe_slug in slug:
                    list_name = maybe_name
                    break
        if not list_name:
            list_name = path.stem.replace("_", " ").title()

        discovered.append(
            ExportedList(
                list_name=list_name,
                url=name_to_url.get(list_name, "manual-export"),
                pulled_at_utc=pulled_at_utc,
                file_path=path,
            )
        )

    if not discovered:
        raise FileNotFoundError(f"No .csv or Excel files found in {download_dir}")

    return discovered


def run_all(download_dir: Path, output_xlsx: Path, manual_login: bool, headless: bool, email: str | None, password: str | None) -> Path:
    exports = download_all_lists(
        download_dir=download_dir,
        headless=headless,
        manual_login=manual_login,
        email=email,
        password=password,
    )
    return build_workbook(exports, output_xlsx)


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "lists": list(IBD_LISTS.keys())}


@app.post("/compile")
def compile_endpoint(request: CompileRequest) -> dict:
    try:
        path = run_all(
            download_dir=Path(request.download_dir),
            output_xlsx=Path(request.output_xlsx),
            manual_login=request.manual_login,
            headless=request.headless,
            email=request.email,
            password=request.password,
        )
        return {"status": "ok", "output_xlsx": str(path)}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/compile-existing-downloads")
def compile_existing_downloads(request: CompileRequest) -> dict:
    try:
        exports = discover_downloads(Path(request.download_dir))
        path = build_workbook(exports, Path(request.output_xlsx))
        return {"status": "ok", "output_xlsx": str(path), "files_used": [str(e.file_path) for e in exports]}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download IBD list exports and compile them into one Excel workbook.")
    sub = parser.add_subparsers(dest="command", required=True)

    run_all_parser = sub.add_parser("run-all", help="Log in to IBD, export the target lists, and build one Excel workbook.")
    run_all_parser.add_argument("--download-dir", default=str(DEFAULT_DOWNLOAD_DIR))
    run_all_parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    run_all_parser.add_argument("--headless", action="store_true")
    run_all_parser.add_argument("--no-manual-login", action="store_true")
    run_all_parser.add_argument("--email", default=os.getenv("IBD_EMAIL"))
    run_all_parser.add_argument("--password", default=os.getenv("IBD_PASSWORD"))

    compile_parser = sub.add_parser("compile-existing", help="Compile already-downloaded IBD exports into one Excel workbook.")
    compile_parser.add_argument("--download-dir", default=str(DEFAULT_DOWNLOAD_DIR))
    compile_parser.add_argument("--output", default=str(DEFAULT_OUTPUT))

    api_parser = sub.add_parser("serve", help="Run a FastAPI server.")
    api_parser.add_argument("--host", default="127.0.0.1")
    api_parser.add_argument("--port", type=int, default=8000)

    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.command == "run-all":
        output = run_all(
            download_dir=Path(args.download_dir),
            output_xlsx=Path(args.output),
            manual_login=not args.no_manual_login,
            headless=args.headless,
            email=args.email,
            password=args.password,
        )
        print(f"Workbook written to: {output}")
        return 0

    if args.command == "compile-existing":
        exports = discover_downloads(Path(args.download_dir))
        output = build_workbook(exports, Path(args.output))
        print(f"Workbook written to: {output}")
        return 0

    if args.command == "serve":
        import uvicorn

        uvicorn.run(app, host=args.host, port=args.port)
        return 0

    print("Unknown command", file=sys.stderr)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
